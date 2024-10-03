import openpyxl

def get_user_data():
    workbook = openpyxl.load_workbook("D:\Interview Learning\pytest_selenium\Data_files/test.xlsx")
    sheet = workbook['Sheet1']
    total_rows = sheet.max_row
    total_col = sheet.max_column
    user_list = []
    for row in range(2, total_rows+1):
        for col in range(1, total_col+1):
            print(sheet.cell(row=row,column=col).value)
            user_list.append(sheet.cell(row=row,column=col).value)
    return user_list


